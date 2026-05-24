import json
import random
import string
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from groq import AsyncGroq
from app.config import get_settings

settings = get_settings()
client = AsyncGroq(api_key=settings.groq_api_key)

# Resolve the Excel path relative to this file: backend/data/support_tickets.xlsx
# savi_agent.py → services/ → app/ → backend/  (3 levels up)
_DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
EXCEL_PATH = _DATA_DIR / "support_tickets.xlsx"

SYSTEM_PROMPT = """You are Savi, the friendly, warm virtual assistant for Savomart stores.
Your goal is to help resolve the customer's issues or escalate their concerns to human support by naturally collecting their details through conversation.

You need to collect the following 4 pieces of information:
1. Customer's Name (e.g., Diya)
2. Contact details (either a 10-digit mobile number or a valid email address)
3. Issue Category (must be exactly one of: 'Order Issue', 'Points / Rewards', 'Coupon Problem', 'Store Feedback', 'Account Help', or 'Other')
4. Detailed Description of the problem (at least 10 characters long)

TONE & PERSONALITY RULES:
- Speak naturally and casually, like a helpful store associate standing next to them. Do not sound robotic, overly formal, or use generic AI boilerplate phrases.
- Welcome the customer warmly and show empathy.
- Guide the conversation to gather the missing details one or two at a time. Do not dump a list of questions all at once or ask the user to fill out a checklist. Be conversational!
- If the customer has already provided some details in their opening message (e.g., their name or problem), acknowledge them, and ask for the remaining ones politely.
- Keep responses friendly, warm, and concise. Always sign off messages with subtle warmth (e.g., "Warmly, Savi", "Have a wonderful day, Savi", "Best, Savi", etc.).

CONFIRMATION & COMPLETION:
- Once you have successfully collected all 4 fields (Name, Contact, Issue Category, and Description), summarize them briefly and ask the customer to confirm if these details are correct.
- If they confirm, or once they say yes, you MUST include the exact phrase "I'll log this for you" in your final response (case-insensitive search will look for this phrase).
- Example final response: "Great, I'll log this for you right away. Our team will look into this and get back to you within 24 hours. Have a great day! Warmly, Savi"
- DO NOT make up or output any ticket ID. The system will handle generating and saving the ticket.
"""


async def get_savi_response(conversation_history: list, user_message: str):
    """
    Generate a streamed response from Savi using the Groq API (llama-3.3-70b-versatile).
    """
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    
    # Add history
    for msg in conversation_history:
        messages.append({"role": msg.role, "content": msg.content})
        
    # Append the new user message
    messages.append({"role": "user", "content": user_message})
    
    # If conversation is reaching its turn limit (20 turns total, including current)
    # We append a system reminder to Savi to log the ticket immediately.
    # Total user messages + assistant messages
    if len(conversation_history) >= 18:
        messages.append({
            "role": "system",
            "content": (
                "CRITICAL: The conversation is ending due to length limits. Do not ask any more questions. "
                "Politely tell the customer that you are escalating this right away, summarize what details you have, "
                "and say the phrase 'I'll log this for you' to submit the ticket immediately."
            )
        })
        
    completion = await client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=messages,
        temperature=0.7,
        stream=True
    )
    
    async for chunk in completion:
        if chunk.choices[0].delta.content is not None:
            yield chunk.choices[0].delta.content


async def extract_ticket_data(conversation_text: str) -> dict:
    """
    Extract structured ticket details from the conversation history using Groq JSON mode.
    """
    extraction_prompt = (
        "You are an expert data extraction assistant. Analyze the following conversation between a customer and Savi (the virtual assistant). "
        "Extract the customer's details into a JSON object with these EXACT keys:\n"
        "- 'name': The customer's name (string, or empty string if not found)\n"
        "- 'contact': The customer's mobile number or email address (string, or empty string if not found)\n"
        "- 'issue_category': Must be exactly one of: 'Order Issue', 'Points / Rewards', 'Coupon Problem', 'Store Feedback', 'Account Help', 'Other'. If not specified, map to the closest one or 'Other'.\n"
        "- 'description': A detailed description of the customer's issue (string, or empty string if not found)\n\n"
        "Return ONLY a valid JSON object. Do not include any explanation or markdown formatting."
    )
    
    response = await client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {"role": "system", "content": extraction_prompt},
            {"role": "user", "content": conversation_text}
        ],
        response_format={"type": "json_object"},
        temperature=0.0
    )
    
    content = response.choices[0].message.content
    try:
        data = json.loads(content)
        return {
            "name": str(data.get("name", "")).strip(),
            "contact": str(data.get("contact", "")).strip(),
            "issue_category": str(data.get("issue_category", "Other")).strip(),
            "description": str(data.get("description", "")).strip()
        }
    except Exception:
        return {
            "name": "",
            "contact": "",
            "issue_category": "Other",
            "description": ""
        }


def save_to_excel(ticket_data: dict) -> str:
    """
    Appends a new support ticket to backend/data/support_tickets.xlsx.
    Creates the file and directories if they do not exist.
    """
    _DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Generate Ticket ID: SAVO-YYYYMMDD-XXXX (date + 4-char random alphanumeric)
    date_str = datetime.now().strftime("%Y%m%d")
    random_chars = "".join(random.choices(string.ascii_uppercase + string.digits, k=4))
    ticket_id = f"SAVO-{date_str}-{random_chars}"

    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if EXCEL_PATH.exists():
        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Support Tickets"
            ws.append(["Ticket ID", "Created At", "Name", "Contact", "Issue Category", "Description", "Status"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Support Tickets"
        ws.append(["Ticket ID", "Created At", "Name", "Contact", "Issue Category", "Description", "Status"])

    ws.append([
        ticket_id,
        created_at,
        ticket_data.get("name", ""),
        ticket_data.get("contact", ""),
        ticket_data.get("issue_category", "Other"),
        ticket_data.get("description", ""),
        "Open"
    ])

    wb.save(EXCEL_PATH)
    wb.close()
    return ticket_id
